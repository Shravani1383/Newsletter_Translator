import os
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import html
import re
import pandas as pd
from itertools import combinations
# import concurrent.futures
import openpyxl
from html.entities import codepoint2name, name2codepoint
import shutil
from pathlib import Path
from zipfile import ZipFile
import streamlit as st
import base64

def read_html_file(html_file_path):
    with open(html_file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()
    return html_content

def excel_to_markdown_table(excel_path):
    print(f"Converting {excel_path} to markdown table format")
    wb = load_workbook(excel_path)
    markdown_tables = ""
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        markdown_tables += f"### Sheet: {sheet_name}\n\n"
        rows = list(sheet.iter_rows(values_only=True))
        
        if rows:
            # Create rows ignoring the first row (headers)
            for row in rows[1:]:
                row_data = []
                for cell in row:
                    if cell is not None:
                        cell_text = str(cell).strip()  # Remove leading and trailing spaces
                        cell_text = re.sub(r'\s+', ' ', cell_text)  # Replace multiple spaces with a single space
                    else:
                        cell_text = ""
                    row_data.append(cell_text)
                markdown_tables += " : ".join(row_data) + " \n"
            markdown_tables += "\n"
    print(markdown_tables)
    return markdown_tables

def convert_to_dict(input_str):
    print("Converting French translation to dict")
    lines = [line.strip() for line in input_str.split("\n") if line.strip()]
    output_dict = {}
    for line in lines:
        if " : " in line:
            key, value = line.split(" : ", 1)
            key = ' '.join(key.split())  # Remove excessive spaces
            value = ' '.join(value.split())  # Remove excessive spaces
            output_dict[key] = value
    print(output_dict)
    return output_dict


def replace_text_in_html(input_file, output_file, replacements):
    def create_pattern(text):
        # Escape the text and allow for &nbsp; in addition to regular spaces and HTML tags
        text = re.escape(text)
        # Replace escaped spaces with regex pattern for spaces, &nbsp;, and any HTML tags
        text = text.replace(r'\ ', r'(\s*|&nbsp;)')
        return text

    with open(input_file, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Sort replacements by the length of the old text in descending order
    sorted_replacements = dict(sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True))
    
    # Replace text in the HTML content (case-insensitive, space-ignoring, ignoring any HTML tags)
    for old, new in sorted_replacements.items():
        pattern = create_pattern(old)
        # Replace &nbsp; with regular space in content before applying regex substitution
        content = content.replace('&nbsp;', ' ')
        content = re.sub(pattern, new, content, flags=re.IGNORECASE)
    
    # Write the updated content to a new file
    with open(output_file, 'w', encoding='utf-8') as file:
        file.write(content)


def extract_columns(file_path, keywords, header_search_rows=10):
    xls = pd.ExcelFile(file_path)
    all_columns = pd.DataFrame()
    seen_columns = set()

    print(f"Sheets found: {xls.sheet_names}")

    for sheet_name in xls.sheet_names:
        if sheet_name == "BALISES":
            print(f"Skipping sheet: {sheet_name}")
            continue
        
        print(f"Processing sheet: {sheet_name}")
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        column_row_index = None
        for i in range(header_search_rows):
            if df.iloc[i].astype(str).str.contains('|'.join(keywords)).any():
                column_row_index = i
                break

        if column_row_index is None:
            print(f"No column names found within the first {header_search_rows} rows in sheet {sheet_name}")
            continue
        
        df.columns = df.iloc[column_row_index].str.strip()
        df = df[column_row_index + 1:]

        print(f"Columns in {sheet_name} (identified row {column_row_index}): {df.columns.tolist()}")

        for col in df.columns:
            if isinstance(col, str):
                for keyword in keywords:
                    if col.startswith(keyword) and col not in seen_columns:
                        print(f"Matched column: {col}")
                        all_columns[col] = df[col]
                        seen_columns.add(col)
                        break

    return all_columns


def create_combinations(df, base_keyword, combination_keywords):
    output_folder = 'processed_excel_files'
    os.makedirs(output_folder, exist_ok=True)
    
    for keyword in combination_keywords:
        if base_keyword in df.columns and keyword in df.columns:
            combined_df = df[[base_keyword, keyword]]
            output_file = os.path.join(output_folder, f'{keyword}.xlsx')
            combined_df.to_excel(output_file, index=False)
            print(f"Created file: {output_file}")


# Regular expression to identify URLs
url_pattern = re.compile(
    r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')

def is_named_entity(text, i):
    if text[i] == '&':
        semicolon_index = text.find(';', i)
        if semicolon_index != -1:
            entity_candidate = text[i:semicolon_index + 1]
            if entity_candidate[1:-1] in name2codepoint:
                return True
    return False

def convert_to_named_entities(text):
    result = []
    i = 0
    while i < len(text):
        if is_named_entity(text, i):
            semicolon_index = text.find(';', i)
            entity_candidate = text[i:semicolon_index + 1]
            result.append(entity_candidate)
            i = semicolon_index + 1
        else:
            code = ord(text[i])
            if code in codepoint2name:
                entity = f"&{codepoint2name[code]};"
                result.append(entity)
            else:
                result.append(text[i])
            i += 1
    return ''.join(result)

def process_excel_file(input_file):
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(input_file, data_only=True)
        print(f"Processing file: {input_file}")
        
        # Iterate through each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing sheet: {sheet_name}")
            
            # Iterate through each row and cell
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        # Process only text cells
                        if isinstance(cell.value, str):
                            original_value = cell.value
                            # Skip URL processing
                            if not url_pattern.search(original_value):
                                cell.value = convert_to_named_entities(cell.value)
                            print(f"Original value in cell {cell.coordinate}: {original_value}")
                            print(f"Converted value in cell {cell.coordinate}: {cell.value}")
                    except Exception as e:
                        print(f"Error processing cell {cell.coordinate}: {e}")
        
        # Save the modified workbook with the same name
        wb.save(input_file)
        print(f"Workbook saved to {input_file}")
    except Exception as e:
        print(f"Failed to process file {input_file}: {e}")

# Iterate over all Excel files in the directory

def process_folder(folder_path, html_template_path):
    html_content = read_html_file(html_template_path)
    
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx"):
            excel_file_path = os.path.join(folder_path, file_name)
            wb = load_workbook(excel_file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                excel_text = excel_to_markdown_table(excel_file_path)
                excel_to_dict=convert_to_dict(excel_text)
                # french_html = extract_french_html(html_content, excel_text)
                # french = extract_french_translation(excel_text)
                html_folder_path = 'html'
                output_html_file = os.path.join(html_folder_path, f"{os.path.splitext(file_name)[0]}.html")
                replace_text_in_html(html_template_path, output_html_file,excel_to_dict)
                print(f"Processed {file_name} - {sheet_name}")



def extract_content(file_content, start_comment, end_comment):
    """Extract content between the specified start and end comments."""
    pattern = re.compile(f'{re.escape(start_comment)}(.*?){re.escape(end_comment)}', re.DOTALL)
    match = pattern.search(file_content)
    extracted_content = match.group(1) if match else ''
    return extracted_content

def replace_content(file_content, new_content, start_comment, end_comment):
    """Replace content between the specified start and end comments with new content."""
    pattern = re.compile(f'({re.escape(start_comment)}).*?({re.escape(end_comment)})', re.DOTALL)
    updated_content = pattern.sub(f'\\1{new_content}\\2', file_content)
    return updated_content

def main(header_footer_dir, html_dir, output_dir):
    if not os.path.exists(header_footer_dir):
        print(f"Error: The directory '{header_footer_dir}' does not exist.")
        return
    if not os.path.exists(html_dir):
        print(f"Error: The directory '{html_dir}' does not exist.")
        return
    
    # Iterate over subdirectories within the header_footer_dir
    for root, dirs, files in os.walk(header_footer_dir):
        for subdir in dirs:
            header_footer_folder = os.path.join(root, subdir)
            if os.path.isdir(header_footer_folder):
                # Construct the corresponding HTML filename
                normalized_subdir = subdir.replace('_', ' ').replace(' ', '')
                found_html_filename = None
                
                # Search for a matching HTML file ignoring extra spaces
                for html_file in os.listdir(html_dir):
                    normalized_html_file = html_file.replace(' ', '')
                    if normalized_html_file == normalized_subdir + '.html':
                        found_html_filename = html_file
                        break
                
                if found_html_filename:
                    html_filepath = os.path.join(html_dir, found_html_filename)
                    index_filepath = os.path.join(header_footer_folder, 'index.html')
                    if os.path.exists(index_filepath):
                        with open(index_filepath, 'r', encoding='utf-8') as file:
                            file1_content = file.read()
                        
                        with open(html_filepath, 'r', encoding='utf-8') as file:
                            file2_content = file.read()

                        # Extract content from index.html
                        header_content = extract_content(file1_content, '<!--Header Code Start-->', '<!--Header Code End-->')
                        footer_content = extract_content(file1_content, '<!--Footer Code Start-->', '<!--Footer Code End-->')

                        # Replace content in the HTML file
                        updated_file2_content = replace_content(file2_content, header_content, '<!--Header Code Start-->', '<!--Header Code End-->')
                        updated_file2_content = replace_content(updated_file2_content, footer_content, '<!--Footer Code Start-->', '<!--Footer Code End-->')

                        # Write the updated content to the output directory
                        output_filepath = os.path.join(output_dir, found_html_filename)
                        os.makedirs(output_dir, exist_ok=True)  # Ensure output directory exists

                        # Debugging output
                        print(f"Writing to {output_filepath}")

                        with open(output_filepath, 'w', encoding='utf-8') as file:
                            file.write(updated_file2_content)
                        print(f"Updated content written to {output_filepath}")
                    else:
                        print(f"Warning: 'index.html' not found in {header_footer_folder}.")
                else:
                    print(f"Warning: Corresponding HTML file for folder '{subdir}' not found in {html_dir}.")


def get_immediate_images_directory(images_directory):
    for root, dirs, files in os.walk(images_directory):
        # Check if the current directory has image files directly within it
        if any(file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')) for file in files):
            return root
    return None

def normalize_name(name):
    # Normalize the name by removing spaces and converting to lowercase
    return name.replace(" ", "").lower()

def organize_html_files(directory_path, images_directory):
    # Ensure the directories exist
    if not os.path.exists(directory_path):
        print(f"Directory '{directory_path}' does not exist.")
        return None
    if not os.path.exists(images_directory):
        print(f"Images directory '{images_directory}' does not exist.")
        return None

    immediate_images_directory = get_immediate_images_directory(images_directory)
    if not immediate_images_directory:
        print(f"No immediate directory containing images found in '{images_directory}'.")
        return None

    # Normalize the name of the immediate images directory for comparison
    normalized_immediate_images_directory = normalize_name(os.path.basename(immediate_images_directory))

    # Iterate through files in the directory
    for filename in os.listdir(directory_path):
        if filename.endswith(".html"):
            file_path = os.path.join(directory_path, filename)

            # Normalize the filename for comparison
            normalized_filename = normalize_name(os.path.splitext(filename)[0])
            
            # Check if there is a corresponding directory name that matches the normalized filename
            matching_directories = [
                d for d in os.listdir(directory_path)
                if os.path.isdir(os.path.join(directory_path, d)) and normalize_name(d) == normalized_filename
            ]

            if matching_directories:
                # Use the first matching directory
                new_directory = os.path.join(directory_path, matching_directories[0])
            else:
                # Create a directory with the normalized filename if no match is found
                new_directory = os.path.join(directory_path, normalized_filename)
                os.makedirs(new_directory, exist_ok=True)

            # Move the HTML file into the new directory
            shutil.move(file_path, new_directory)

            # Copy the immediate 'Images' directory into the new directory
            new_images_directory = os.path.join(new_directory, "images")
            shutil.copytree(immediate_images_directory, new_images_directory, symlinks=True)

            print(f"Moved '{filename}' and copied images from '{immediate_images_directory}' to '{new_directory}'.")

    return directory_path  # Return the organized files directory

# Create a file uploader widget


# Example usage
st.title("Upload Directories and Files")




# Function to upload and extract zip file
def upload_and_extract_zip(label):
    uploaded_file = st.file_uploader(label, type="zip")
    if uploaded_file:
        zip_path = Path(uploaded_file.name)
        with open(zip_path, 'wb') as f:
            f.write(uploaded_file.read())
        
        # Extract zip file
        extracted_path = Path(label.replace(" ", "_").replace(".zip", ""))
        with ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extracted_path)
        
        return extracted_path
    return None

# Function to upload a single file
def upload_file(label, file_types):
    uploaded_file = st.file_uploader(label, type=file_types)
    if uploaded_file:
        file_path = Path(uploaded_file.name)
        with open(file_path, 'wb') as f:
            f.write(uploaded_file.read())
        return file_path
    return None

def remove_dir(directory_path):
    """
    Removes a directory and all its contents.

    :param directory_path: Path to the directory to be removed.
    """
    if os.path.exists(directory_path):
        shutil.rmtree(directory_path)
        print(f"Directory '{directory_path}' has been removed.")
    else:
        print(f"Directory '{directory_path}' does not exist.")
def clear_directory(directory_path):
    """
    Clears all files and subdirectories within the specified directory.

    Parameters:
    directory_path (str): The path to the directory to be cleared.
    """
    # Check if the directory exists
    if os.path.exists(directory_path):
        # Iterate over all the files and directories within the specified directory
        for item in os.listdir(directory_path):
            item_path = os.path.join(directory_path, item)
            # If it's a directory, remove it and its contents
            if os.path.isdir(item_path):
                shutil.rmtree(item_path)
            # If it's a file, remove it
            elif os.path.isfile(item_path) or os.path.islink(item_path):
                os.unlink(item_path)
    else:
        print(f"The directory {directory_path} does not exist.")


def recreate_directory(directory):
    """
    Recreates the specified directory. If it already exists, it is removed first.
    
    Args:
    - directory (str): The path of the directory to recreate.
    """
    if os.path.exists(directory):
        shutil.rmtree(directory)  # Remove the directory and its contents
        print(f"Directory removed: {directory}")
    
    os.makedirs(directory)  # Recreate the directory
    print(f"Directory created: {directory}")

# Define the path to the input Excel file
input_excel_file_path = upload_file("Upload Excel File", ["xlsx"])
output_file=input_excel_file_path
# output_file='final_encoded_excel_file.xlsx'
# if input_excel_file_path is not None:
#     html_encoded_excel = convert_excel_text_to_html_entities(input_excel_file_path, output_file)
# else:
#     st.error("Please upload a file.")
# Keywords to search for in column headers
# keywords = ['FR','NO']
# keywords=['FR','SE']
# Process the generated Excel files in the output folder
output_folder_path = 'processed_excel_files'
recreate_directory(output_folder_path)
# remove_dir(output_folder_path)

html_template_path = upload_file("Upload HTML Template File", ["html"])
header_footer_dir = upload_and_extract_zip("Header and Footer")
html_dir = 'html'
recreate_directory(html_dir)
# remove_dir(html_dir)

output_dir = 'Translated_Files'
recreate_directory(output_dir)
# remove_dir(output_dir)

images_dir = upload_and_extract_zip("images")
base_keyword=st.text_input('Enter your Master Language (e.g., "FR")', '')
# combination_keywords = [keyword for keyword in keywords if keyword != base_keyword]

uploaded_file = st.file_uploader("Upload a text file with all languages", type="txt")

if uploaded_file:
    # Read the content of the uploaded file
    content = uploaded_file.read().decode('utf-8')
    
    # Process the content to create a list of keywords
    keywords = [line.strip() for line in content.splitlines() if line.strip()]
    
    # Display the uploaded keywords
    st.write('Uploaded keywords:', keywords)
    
    # Create a new list excluding the base keyword
    combination_keywords = [keyword for keyword in keywords if keyword != base_keyword]
    
    # Display the combination keywords
    st.write('Combination keywords (excluding base keyword):', combination_keywords)
else:
    st.write('Please upload a text file with keywords.')

# Extract columns
if st.button("Process Files"):
    if input_excel_file_path and html_template_path and header_footer_dir:
        st.write("Processing files...")
        # output_file = 'output_converted.xlsx'
        # html_encoded_excel=convert_excel_text_to_html_entities(input_excel_file_path,output_file)
        all_columns = extract_columns(output_file, keywords)
        # base_keyword = 'FR'
        
        # combination_keywords = ['SE']
        # combination_keywords = ['WW', 'NL','ES', 'DE','IT', 'NO', 'SE', 'DK','ME_AR','ME_EN']
       
        create_combinations(all_columns, base_keyword,combination_keywords)
        files_processed = 0
        for filename in os.listdir(output_folder_path):
            if filename.endswith(".xlsx"):
                input_path = os.path.join(output_folder_path, filename)
                process_excel_file(input_path)
                files_processed += 1

        print(f"Total files processed: {files_processed}")


        process_folder(output_folder_path, html_template_path)

        # os.makedirs(output_dir, exist_ok=True)
        main(header_footer_dir, html_dir, output_dir)
        
        Final_output_dir = organize_html_files(output_dir, images_dir)
        
        if Final_output_dir:
            def zip_directory(directory_path, zip_path):
                with ZipFile(zip_path, 'w') as zipf:
                    for root, _, files in os.walk(directory_path):
                        for file in files:
                            file_path = os.path.join(root, file)
                            zipf.write(file_path, os.path.relpath(file_path, os.path.dirname(directory_path)))

    # Zip the output directory
            zip_path = f"{Final_output_dir}.zip"
            zip_directory(Final_output_dir, zip_path)

            with open(zip_path, 'rb') as f:
                # Read the zip file content
                zip_data = f.read()

                # Custom CSS to align the button to the right
                st.markdown("""
                    <style>
                    .download-button-wrapper {
                        display: flex;
                        justify-content: flex-end;
                    }
                    </style>
                    """, 
                    unsafe_allow_html=True)

                # Encode zip data to base64
                zip_data_b64 = base64.b64encode(zip_data).decode()

                # Place the button inside a div with the custom style
                st.markdown(f"""
                    <div class="download-button-wrapper">
                        <a href="data:application/zip;base64,{zip_data_b64}" download="{os.path.basename(zip_path)}" class="stDownloadButton">
                            <button>Download Translated Files</button>
                        </a>
                    </div>""", unsafe_allow_html=True)
                
                # empty_directory(Final_output_dir)

        else:
            st.write("No files were organized or final output directory not found.")



